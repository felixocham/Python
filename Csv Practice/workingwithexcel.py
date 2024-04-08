# in this program I use the openpyxl to analyse inventory in a spreadsheet.
import openpyxl

inventory_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inventory_file['Sheet 1']

product_per_supplier = {}
inventory_value = {}
products_under_10 = {}

for product_row in range (2,product_list.max_row+1):
    supplier_name = product_list.cell(product_row,4).value
    inventory_price = product_list.cell(product_row, 3).value
    inventory_qty = product_list.cell(product_row, 2).value
    product_name = product_list.cell(product_row,1).value
    inv_price = product_list.cell(product_row,5)

    # calculation number of inventory per supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        product_per_supplier[supplier_name] = 1

    # calculation total inventory value per supplier
    if supplier_name in inventory_value:
        current_inv_value = inventory_value.get(supplier_name)
        inventory_value[supplier_name] = current_inv_value + (inventory_qty*inventory_qty)
    else:
        inventory_value[supplier_name] = (inventory_qty*inventory_qty)

    # product with inventory under 10
    if inventory_qty < 10:
        products_under_10[int(product_name)] = int(inventory_qty)

    # adding inventory value for each column
    inv_price.value = (inventory_qty*inventory_qty)


print(f' product num: inventory qty:\n {products_under_10}')
print(f'company name: inventory quantity: \n {product_per_supplier}')
print(f'company name: inventory value: \n {inventory_value}')


inventory_file.save('Updated Inventory file.xls')
